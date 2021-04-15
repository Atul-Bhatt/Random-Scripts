from openpyxl import Workbook
from openpyxl import load_workbook
import os

wb = load_workbook('exceldata.xlsx')

active_sheet = wb.active

for row in active_sheet.iter_rows():
    file_with_path = os.path.join("C:\\Users\\I516396\\Desktop\\experimental", f'{row[0].value} - {row[1].value}.txt')
    with open(file_with_path, 'w') as file:
        file.write(row[2].value)
